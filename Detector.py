import os
import cv2 #Face Recognition Library
from time import sleep
from PIL import Image
from datetime import datetime
import openpyxl  #Excel Library

def main_app():
    face_cascade = cv2.CascadeClassifier('./data/haarcascade_frontalface_default.xml')
    cap = cv2.VideoCapture(0)
    pred = 0

    # Get the list of classifier files in the /data/classifiers directory
    classifier_dir = './data/classifiers'
    classifier_files = os.listdir(classifier_dir)
    classifier_files = [file for file in classifier_files if file.endswith('_classifier.xml')]

    # Generate file name with today's date
    today_date = datetime.now().strftime("%Y-%m-%d")
    file_name = f"SECD2052 - {today_date}.xlsx"
    workbook_path = os.path.join('./data', file_name)

    # Create or load the Excel workbook
    if os.path.exists(workbook_path):
        workbook = openpyxl.load_workbook(workbook_path)
    else:
        workbook = openpyxl.Workbook()

    sheet = workbook.active
    if 'User Data' not in workbook.sheetnames:
        sheet.title = 'User Data'
        sheet.append(['Name', 'Timestamp'])

    # Set to keep track of written names
    written_names = set()


    while True: #loop1 to open camera and scan for faces

        ret, frame = cap.read()
        gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
        faces = face_cascade.detectMultiScale(gray, 1.3, 5)

        highest_confidence = 0
        predicted_name = "UnknownFace"

        for (x, y, w, h) in faces: #Loop2 to find classifier file
            roi_gray = gray[y:y+h, x:x+w]

            # Loop through each classifier file
            for classifier_file in classifier_files: #Loop3 to match face profile
                recognizer = cv2.face.LBPHFaceRecognizer_create()
                recognizer.read(os.path.join(classifier_dir, classifier_file))

                id, confidence = recognizer.predict(roi_gray)
                confidence = 100 - int(confidence)

                if confidence > highest_confidence:
                    highest_confidence = confidence
                    predicted_name = os.path.splitext(classifier_file)[0].split('_')[0].upper() 

            # Display the predicted name and bounding box on the frame
            font = cv2.FONT_HERSHEY_PLAIN
            color = (0, 255, 0) if highest_confidence > 50 else (0, 0, 255)
            frame = cv2.rectangle(frame, (x, y), (x + w, y + h), color, 2)
            frame = cv2.putText(frame, predicted_name, (x, y-4), font, 1, color, 1, cv2.LINE_AA)

        # Save user data to Excel file if highest confidence is greater than 0 and name is not written before
        if highest_confidence > 0 and predicted_name not in written_names:
            timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            sheet.append([predicted_name, timestamp])
            written_names.add(predicted_name)
            workbook.save(workbook_path)

        cv2.imshow("image", frame)

        if cv2.waitKey(20) & 0xFF == ord('q'):
            break


    cap.release()
    cv2.destroyAllWindows()
